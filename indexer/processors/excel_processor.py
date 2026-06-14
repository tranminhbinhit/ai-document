import logging
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

def extract_text_from_excel(file_path: str) -> str:
    """Extract text from Excel file"""
    try:
        workbook = load_workbook(file_path, data_only=True)
        text = ""
        
        for sheet in workbook.worksheets:
            text += f"Sheet: {sheet.title}\n"
            for row in sheet.iter_rows(values_only=True):
                row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    text += row_text + "\n"
            text += "\n"
        
        return text.strip()
    except Exception as e:
        logger.error(f"Error extracting text from Excel: {e}")
        raise
