"""
MCP Filesystem Server - Đọc tài liệu và phân tích file
Hỗ trợ: PDF, DOCX, XLSX, TXT
"""

import os
import json
from pathlib import Path
from typing import Any

try:
    from mcp.server import Server
    from mcp.types import Tool, TextContent
except ImportError:
    print("Install: pip install mcp")
    exit(1)

# Document parsers
try:
    import PyPDF2
    import docx
    import openpyxl
except ImportError:
    print("Install: pip install PyPDF2 python-docx openpyxl")
    exit(1)

app = Server("mcp-filesystem-server")

WORKSPACE_ROOT = os.getenv("WORKSPACE_ROOT", ".")
ALLOWED_DIRS = os.getenv("ALLOWED_DIRECTORIES", WORKSPACE_ROOT).split(",")


def is_path_allowed(path: str) -> bool:
    """Kiểm tra path có nằm trong thư mục được phép"""
    abs_path = os.path.abspath(path)
    return any(abs_path.startswith(os.path.abspath(d)) for d in ALLOWED_DIRS)


@app.list_tools()
async def list_tools() -> list[Tool]:
    return [
        Tool(
            name="list_directory",
            description="Liệt kê files và folders trong thư mục",
            inputSchema={
                "type": "object",
                "properties": {
                    "path": {
                        "type": "string",
                        "description": "Đường dẫn thư mục (relative hoặc absolute)"
                    }
                },
                "required": ["path"]
            }
        ),
        Tool(
            name="read_text_file",
            description="Đọc nội dung file text (txt, md, json, etc.)",
            inputSchema={
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "Đường dẫn file"}
                },
                "required": ["path"]
            }
        ),
        Tool(
            name="parse_pdf",
            description="Đọc và extract text từ file PDF",
            inputSchema={
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "Đường dẫn file PDF"}
                },
                "required": ["path"]
            }
        ),
        Tool(
            name="parse_docx",
            description="Đọc và extract text từ file Word (DOCX)",
            inputSchema={
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "Đường dẫn file DOCX"}
                },
                "required": ["path"]
            }
        ),
        Tool(
            name="parse_xlsx",
            description="Đọc và extract data từ file Excel (XLSX)",
            inputSchema={
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "Đường dẫn file XLSX"},
                    "sheet_name": {"type": "string", "description": "Tên sheet (optional)"}
                },
                "required": ["path"]
            }
        )
    ]


@app.call_tool()
async def call_tool(name: str, arguments: Any) -> list[TextContent]:
    try:
        if name == "list_directory":
            return await handle_list_directory(arguments)
        elif name == "read_text_file":
            return await handle_read_text_file(arguments)
        elif name == "parse_pdf":
            return await handle_parse_pdf(arguments)
        elif name == "parse_docx":
            return await handle_parse_docx(arguments)
        elif name == "parse_xlsx":
            return await handle_parse_xlsx(arguments)
        else:
            return [TextContent(type="text", text=f"Unknown tool: {name}")]
    except Exception as e:
        return [TextContent(type="text", text=f"Error: {str(e)}")]


async def handle_list_directory(args: dict) -> list[TextContent]:
    path = args["path"]
    if not is_path_allowed(path):
        return [TextContent(type="text", text="Error: Path not allowed")]
    
    items = []
    for item in Path(path).iterdir():
        items.append({
            "name": item.name,
            "type": "directory" if item.is_dir() else "file",
            "size": item.stat().st_size if item.is_file() else None
        })
    
    return [TextContent(type="text", text=json.dumps(items, indent=2, ensure_ascii=False))]


async def handle_read_text_file(args: dict) -> list[TextContent]:
    path = args["path"]
    if not is_path_allowed(path):
        return [TextContent(type="text", text="Error: Path not allowed")]
    
    with open(path, "r", encoding="utf-8") as f:
        content = f.read()
    
    return [TextContent(type="text", text=content)]


async def handle_parse_pdf(args: dict) -> list[TextContent]:
    path = args["path"]
    if not is_path_allowed(path):
        return [TextContent(type="text", text="Error: Path not allowed")]
    
    text_content = []
    with open(path, "rb") as f:
        pdf_reader = PyPDF2.PdfReader(f)
        for page_num, page in enumerate(pdf_reader.pages, 1):
            text = page.extract_text()
            text_content.append(f"--- Page {page_num} ---\n{text}\n")
    
    return [TextContent(type="text", text="\n".join(text_content))]


async def handle_parse_docx(args: dict) -> list[TextContent]:
    path = args["path"]
    if not is_path_allowed(path):
        return [TextContent(type="text", text="Error: Path not allowed")]
    
    doc = docx.Document(path)
    paragraphs = [para.text for para in doc.paragraphs if para.text.strip()]
    
    return [TextContent(type="text", text="\n\n".join(paragraphs))]


async def handle_parse_xlsx(args: dict) -> list[TextContent]:
    path = args["path"]
    if not is_path_allowed(path):
        return [TextContent(type="text", text="Error: Path not allowed")]
    
    wb = openpyxl.load_workbook(path, read_only=True)
    sheet_name = args.get("sheet_name") or wb.sheetnames[0]
    sheet = wb[sheet_name]
    
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append([str(cell) if cell is not None else "" for cell in row])
    
    return [TextContent(type="text", text=json.dumps(data, indent=2, ensure_ascii=False))]


if __name__ == "__main__":
    import asyncio
    import sys
    from mcp.server.stdio import stdio_server
    
    # Fix UTF-8 encoding for Windows console
    if sys.platform == "win32":
        import io
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
    
    async def main():
        async with stdio_server() as (read_stream, write_stream):
            await app.run(read_stream, write_stream, app.create_initialization_options())
    
    asyncio.run(main())
